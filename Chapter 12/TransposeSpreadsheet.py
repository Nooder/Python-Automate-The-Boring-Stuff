#! python3
# Chapter 12 Project - Transpose rows/cols in a spreadsheet

import openpyxl, sys, logging

if len(sys.argv) < 2:
    print("USAGE: TransposeSpreadsheet.py <filename>")
    sys.exit()

# Setup
filename = sys.argv[1]
wb = openpyxl.load_workbook(filename)
sheet = wb.active
newWb = openpyxl.Workbook()
newSheet = newWb.active

# Transpose the rows/cols
for rowNum in range(1, sheet.max_row + 1):
    for colNum in range(1, sheet.max_column + 1):
        newSheet.cell(colNum, rowNum).value = sheet.cell(rowNum, colNum).value

newWb.save(filename + "_TRANSPOSED.xlsx")