#! python3
# Chapter 12 Project - Read a spreadsheet and each row to a new text file
# USAGE: XlsxToTxt.py <filename.xlsx>

import openpyxl, sys, logging

#Setup
if len(sys.argv) < 2:
    print("USAGE: XlsxToTxt.py <filename.xlsx>")

filename = sys.argv[1]
wb = openpyxl.load_workbook(filename)
sheet = wb.active

# Loop through each row in the spreadsheet and write to a new text file for each
for rowNum in range(1, sheet.max_row + 1):
    file = open("XlsxToTxt" + str(rowNum) + ".txt", "w")
    for colNum in range(1, sheet.max_column + 1):
        try:
            file.write(sheet.cell(rowNum, colNum).value)
        except TypeError:
            file.write("\n")
    file.close()

print("Done.")