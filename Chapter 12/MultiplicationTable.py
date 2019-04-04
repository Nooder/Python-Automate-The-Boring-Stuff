#! python3
# Chapter 12 - Generate a multiplication table in Excel
# USAGE: MultiplicationTable.py <number>       # Number to go up to

import openpyxl, sys, logging
from openpyxl.styles import Font

if len(sys.argv) < 2:
    print("USAGE: MultiplicationTable.py <number>")
    sys.exit()

# Setup
endNumber = int(sys.argv[1])
wb = openpyxl.Workbook()
sheet = wb.active
headerFont = Font(bold=True)

# Fill in the headers
for rowNum in range(2, endNumber + 2):
    sheet.cell(row=rowNum, column=1).value = rowNum - 1
    sheet.cell(row=rowNum, column=1).font = headerFont

for colNum in range(2, endNumber + 2):
    sheet.cell(row=1, column=colNum).value = colNum - 1
    sheet.cell(row=1, column=colNum).font = headerFont

# Fill in the table
for rowNum in range(2, endNumber + 2):
    for colNum in range(2, endNumber + 2):
        sheet.cell(row=rowNum, column=colNum).value = (rowNum - 1) * (colNum - 1)

wb.save("multiplicationTable.xlsx")