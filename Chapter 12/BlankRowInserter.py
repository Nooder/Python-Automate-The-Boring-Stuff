#! python3
# Chapter 12 Project - Insert blank rows in to a spreadsheet
# USAGE: BlackRowInserter.py <N> <M> <filename>
# N = Row to start inserting blank  M = # of blanks to insert

import openpyxl, sys, logging

# Setup
if len(sys.argv) < 4:
    print("USAGE: BlackRowInserter.py <N> <M> <filename>")
    sys.exit()

n = int(sys.argv[1])
m = int(sys.argv[2])
filename = sys.argv[3]

wb = openpyxl.load_workbook(filename)
sheet = wb.active
newWb = openpyxl.Workbook()
newSheet = newWb.active

# Copy rows 1 to N
for rowNum in range(1, n + 1):
    for colNum in range(1, sheet.max_column + 1):
        newSheet.cell(rowNum, colNum).value = sheet.cell(rowNum, colNum).value

# Copy rows N + M to END
for rowNum in range(n + m, sheet.max_row + m):
    for colNum in range(1, sheet.max_column + 1):
        newSheet.cell(rowNum, colNum).value = sheet.cell(rowNum - n, colNum).value

newWb.save(filename + "_BLANKED.xlsx")