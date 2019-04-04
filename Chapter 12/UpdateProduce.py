#! python3
# Chapter 12 - Corrects costs in the produce spreadsheet

import openpyxl

wb = openpyxl.load_workbook('produceSales.xlsx')
sheet = wb.get_sheet_by_name('Sheet')

# The produce type and their updated prices
PRICE_UPDATES = {
    'Garlic' : 3.07,
    'Celery' : 1.19,
    'Lemon' : 1.27
}

# Loop through the rows and update the prices
for rowNum in range(2, sheet.max_row):  # Skip the first row
    produceName = sheet.cell(row=rowNum, column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum, column=1).value = PRICE_UPDATES[produceName]

wb.save('updatedProduceSales.xlsx')