#! python3
# Chapter 12 Project - Read multiple txt files in to a spreadsheet
# USAGE: TxtToXlsx.py <resultFile.xlsx> <txtFile1.txt> ... <txtFileN.txt>

import openpyxl, sys, logging

# Setup
if len(sys.argv) < 3:
    print("USAGE: TxtToXlsx.py <resultFile.xlsx> <txtFile1.txt> ... <txtFileN.txt>")
    sys.exit()

resultFilename = sys.argv[1]
wb = openpyxl.Workbook()
sheet = wb.active
textFiles = []
# Store each text file arg in a list
for arg in range(2, len(sys.argv)):
    textFiles.append(sys.argv[arg])

# Read in files and add them to the spreadsheet
for textFile in textFiles:
    fileName = open(textFile)
    lines = fileName.readlines()
    # Write each line in the text file to the spreadsheet in this row
    sheet.cell(row=textFiles.index(textFile) + 1, column=1).value = textFile + ": "
    for line in lines:
        sheet.cell(row=textFiles.index(textFile) + 1, column=lines.index(line) + 2).value = line

wb.save(resultFilename)