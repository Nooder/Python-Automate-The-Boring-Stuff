#! python3
# Chapter 14 Project - Convert all xlsx files in cwd to csv files

import openpyxl, sys, os, csv, logging

# Find xlsx files
xlsxFiles = []
os.makedirs('Converted Xlsx-Csv Files', exist_ok=True)
for filename in os.listdir('.'):
    if filename.endswith(".xlsx"):
        xlsxFiles.append(filename)

for excelFile in xlsxFiles:
    print("Reading %s..." % excelFile)
    wb = openpyxl.load_workbook(excelFile)
    # Loop through the sheets in the current wb and add them to a csv per sheet
    for sheetName in wb.sheetnames:
        sheet = wb[sheetName]
        csvFile = open(os.path.join('Converted Xlsx-Csv Files', excelFile[:-4] + "_" + sheetName + ".csv"), 'w')
        csvWriter = csv.writer(csvFile)
        # Loop through the rows in the sheet and add them to the csv
        for rowNum in range(1, sheet.max_row + 1):
            rowData = []    # Append each cell from this row to this list
            # Loop through each cell in the row
            for colNum in range(1, sheet.max_column + 1):
                rowData.append(sheet.cell(rowNum, colNum).value)
            csvWriter.writerow(rowData)
        print("Writing %s..." % (excelFile[:-4] + "_" + sheetName + ".csv"))
        csvFile.close()

print("Done.")